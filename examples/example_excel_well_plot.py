"""
Example: Plot well production and injection history from an Excel workbook.

The workbook is expected to have:
- Column B: date
- Column C: well name
- Columns D to I: Qgm, Qwm, Qom, Qclm, Qwim, Qgim

At startup, the user selects the workbook file and the GUI opens with a well
selector that can be changed at any time.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np

try:
    import tkinter as tk
    from tkinter import filedialog, ttk
except ImportError:
    tk = None
    filedialog = None
    ttk = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - import guard for runtime use
    raise SystemExit(
        "openpyxl is required to read Excel files. Install dependencies with "
        "'pip install -r requirements.txt'."
    ) from exc


VARIABLES = [
    ("Qgm", "Gas production"),
    ("Qwm", "Water production"),
    ("Qom", "Oil production"),
    ("Qclm", "Condensate production"),
    ("Qwim", "Water injection"),
    ("Qgim", "Gas injection"),
]

FIT_VARIABLE_MAP = {
    "Qg": "Qgm",
    "Qo": "Qom",
    "Qw": "Qwm",
}

FIT_MODEL_OPTIONS = ["Auto", "Exponential", "Hyperbolic", "Harmonic", "Asymptotic Exponential"]


def ask_workbook_path() -> Path:
    """Ask the user for the workbook path at application startup."""
    if tk is not None and filedialog is not None:
        root = tk.Tk()
        root.withdraw()
        selected = filedialog.askopenfilename(
            title="Select Excel workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
            initialdir=str(Path(__file__).resolve().parent.parent),
        )
        root.destroy()

        if selected:
            return Path(selected)

    while True:
        typed_path = input("Enter the workbook path (.xlsx/.xlsm/.xls): ").strip().strip('"')
        if not typed_path:
            print("Path cannot be empty. Try again.")
            continue

        candidate = Path(typed_path)
        if candidate.exists() and candidate.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            return candidate

        print("Invalid workbook path. Try again.")


def normalize_well_name(value: object) -> str:
    """Convert workbook cell values to a clean well name string."""
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value: object) -> datetime:
    """Validate and normalize date values from the workbook."""
    if isinstance(value, datetime):
        return value
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, str):
        for parser in (datetime.fromisoformat,):
            try:
                return parser(value)
            except ValueError:
                continue
    raise ValueError(f"Unsupported date value: {value!r}")


def read_workbook(filepath: Path, sheet_name: str | None = None) -> dict[str, dict[str, list]]:
    """Read the workbook and group rows by well name."""
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    wells: dict[str, dict[str, list]] = defaultdict(lambda: {"date": [], **{key: [] for key, _ in VARIABLES}})

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 9:
            continue

        date_value = row[1]
        well_name = normalize_well_name(row[2])
        if not date_value or not well_name:
            continue

        try:
            date = parse_date(date_value)
        except ValueError:
            continue

        well_bucket = wells[well_name]
        well_bucket["date"].append(date)

        for index, (variable_name, _) in enumerate(VARIABLES, start=3):
            value = row[index]
            if value is None or value == "":
                well_bucket[variable_name].append(None)
            else:
                well_bucket[variable_name].append(float(value))

    if not wells:
        raise ValueError("No wells were found in the workbook.")

    return wells


def get_sorted_well_series(well_data: dict[str, list]) -> list[tuple]:
    """Return rows sorted by date for the selected well."""
    return sorted(
        zip(well_data["date"], *[well_data[var] for var, _ in VARIABLES]),
        key=lambda item: item[0],
    )


def fit_exponential(t_years: np.ndarray, q_values: np.ndarray) -> dict[str, Any] | None:
    """Fit exponential decline model: q = qi * exp(-Di*t)."""
    if len(t_years) < 2:
        return None

    ln_q = np.log(q_values)
    slope, intercept = np.polyfit(t_years, ln_q, 1)
    di = -slope
    qi = float(np.exp(intercept))
    if qi <= 0:
        return None

    q_pred = qi * np.exp(-di * t_years)
    rmse = float(np.sqrt(np.mean((q_values - q_pred) ** 2)))
    return {
        "model": "Exponential",
        "qi": float(qi),
        "Di": float(di),
        "b": None,
        "q_pred": q_pred,
        "rmse": rmse,
    }


def fit_harmonic(t_years: np.ndarray, q_values: np.ndarray) -> dict[str, Any] | None:
    """Fit harmonic decline model: q = qi / (1 + Di*t)."""
    if len(t_years) < 2:
        return None

    inv_q = 1.0 / q_values
    slope, intercept = np.polyfit(t_years, inv_q, 1)
    if intercept <= 0:
        return None

    qi = 1.0 / intercept
    di = slope * qi
    q_pred = qi / (1.0 + di * t_years)
    rmse = float(np.sqrt(np.mean((q_values - q_pred) ** 2)))
    return {
        "model": "Harmonic",
        "qi": float(qi),
        "Di": float(di),
        "b": 1.0,
        "q_pred": q_pred,
        "rmse": rmse,
    }


def fit_hyperbolic(t_years: np.ndarray, q_values: np.ndarray) -> dict[str, Any] | None:
    """Fit hyperbolic decline model by searching b in (0, 1)."""
    if len(t_years) < 2:
        return None

    best_fit: dict[str, Any] | None = None
    for b in np.linspace(0.05, 0.95, 91):
        y = q_values ** (-b)
        slope, intercept = np.polyfit(t_years, y, 1)
        if intercept <= 0 or b <= 0:
            continue

        qi = intercept ** (-1.0 / b)
        di = slope / (b * intercept)
        denominator = (1.0 + b * di * t_years) ** (1.0 / b)
        if np.any(denominator <= 0):
            continue

        q_pred = qi / denominator
        rmse = float(np.sqrt(np.mean((q_values - q_pred) ** 2)))
        candidate = {
            "model": "Hyperbolic",
            "qi": float(qi),
            "Di": float(di),
            "b": float(b),
            "q_pred": q_pred,
            "rmse": rmse,
        }

        if best_fit is None or candidate["rmse"] < best_fit["rmse"]:
            best_fit = candidate

    return best_fit


def build_dense_time_axis(dates: list[datetime], n_points: int = 250) -> tuple[list[datetime], np.ndarray]:
    """Build a dense time axis (dates and years) for smooth fit visualization."""
    if not dates:
        return [], np.array([], dtype=float)

    t0 = dates[0]
    t_end_years = max(0.0, (dates[-1] - t0).days / 365.25)
    dense_t_years = np.linspace(0.0, t_end_years, max(2, n_points))
    dense_dates = [t0 + timedelta(days=float(t_year * 365.25)) for t_year in dense_t_years]
    return dense_dates, dense_t_years


def fit_asymptotic_exponential(t_years: np.ndarray, q_values: np.ndarray) -> dict[str, Any] | None:
    """Fit asymptotic exponential model: y = 1 - exp(-Di*t)."""
    if len(t_years) < 2:
        return None

    q_min = float(np.min(q_values))
    q_max = float(np.max(q_values))
    q_range = q_max - q_min
    if q_range > 0 and (q_min < 0.0 or q_max > 1.0):
        scaled_values = (q_values - q_min) / q_range
        normalized = True
    else:
        scaled_values = np.clip(q_values, 0.0, 0.999999)
        q_min = 0.0
        q_max = 1.0
        normalized = False

    valid_mask = (t_years > 0) & (scaled_values >= 0.0) & (scaled_values < 1.0)
    if np.count_nonzero(valid_mask) < 2:
        return None

    z_values = -np.log1p(-np.clip(scaled_values[valid_mask], 0.0, 0.999999))
    slope = float(np.dot(t_years[valid_mask], z_values) / np.dot(t_years[valid_mask], t_years[valid_mask]))
    di = max(slope, 0.0)

    q_pred_scaled = 1.0 - np.exp(-di * t_years)
    if normalized:
        q_pred = q_min + q_pred_scaled * q_range
    else:
        q_pred = q_pred_scaled

    rmse = float(np.sqrt(np.mean((q_values - q_pred) ** 2)))
    return {
        "model": "Asymptotic Exponential",
        "qi": None,
        "Di": float(di),
        "b": None,
        "q_pred": q_pred,
        "rmse": rmse,
        "scale_min": q_min,
        "scale_max": q_max,
        "normalized": normalized,
    }


def fit_decline_model(
    dates: list[datetime],
    values: list[float],
    forced_model: str | None = None,
) -> dict[str, Any]:
    """Fit decline model and optionally force one specific model.

    Auto mode only compares traditional engineering decline models.
    """
    if len(dates) < 2:
        raise ValueError("At least two valid points are required to fit decline models.")

    t0 = dates[0]
    t_years = np.array([(date - t0).days / 365.25 for date in dates], dtype=float)
    q_values = np.array(values, dtype=float)

    model_fitters = {
        "Exponential": lambda: fit_exponential(t_years, q_values),
        "Hyperbolic": lambda: fit_hyperbolic(t_years, q_values),
        "Harmonic": lambda: fit_harmonic(t_years, q_values),
        "Asymptotic Exponential": lambda: fit_asymptotic_exponential(t_years, q_values),
    }

    if forced_model and forced_model != "Auto":
        if forced_model not in model_fitters:
            raise ValueError(f"Unsupported forced model: {forced_model}")
        best = model_fitters[forced_model]()
        if best is None:
            raise ValueError(f"Could not fit model {forced_model} for the selected interval.")
    else:
        # Auto uses only traditional decline models.
        candidates = [
            model_fitters["Exponential"](),
            model_fitters["Harmonic"](),
            model_fitters["Hyperbolic"](),
        ]
        valid_candidates = [candidate for candidate in candidates if candidate is not None]
        if not valid_candidates:
            raise ValueError("Could not fit any decline model for the selected interval.")
        best = min(valid_candidates, key=lambda candidate: candidate["rmse"])
    fit_dates = dates
    fit_values = [float(value) for value in best["q_pred"]]

    if best["model"] == "Asymptotic Exponential":
        dense_dates, dense_t_years = build_dense_time_axis(dates)
        fit_dates = dense_dates
        q_min = float(best.get("scale_min", 0.0))
        q_max = float(best.get("scale_max", 1.0))
        q_range = q_max - q_min
        dense_scaled = 1.0 - np.exp(-float(best["Di"]) * dense_t_years)
        fit_values = [float(value) for value in (q_min + dense_scaled * q_range)]
    return {
        "model": best["model"],
        "qi": best["qi"],
        "Di": best["Di"],
        "b": best["b"],
        "rmse": best["rmse"],
        "scale_min": best.get("scale_min"),
        "scale_max": best.get("scale_max"),
        "normalized": best.get("normalized", False),
        "fit_dates": fit_dates,
        "fit_values": fit_values,
    }


def calculate_decline_curve(
    model_name: str,
    qi: float,
    di: float | None,
    b_value: float | None,
    dates: list[datetime],
) -> list[float]:
    """Calculate q(t) values for a decline model over provided dates."""
    if not dates:
        return []

    t0 = dates[0]
    t_years = np.array([(date - t0).days / 365.25 for date in dates], dtype=float)

    if model_name == "Exponential":
        if di is None:
            raise ValueError("Missing Di for exponential model.")
        return [float(value) for value in qi * np.exp(-di * t_years)]
    if model_name == "Harmonic":
        if di is None:
            raise ValueError("Missing Di for harmonic model.")
        return [float(value) for value in qi / (1.0 + di * t_years)]
    if model_name == "Hyperbolic":
        if di is None:
            raise ValueError("Missing Di for hyperbolic model.")
        if b_value is None or b_value <= 0:
            raise ValueError("Invalid b value for hyperbolic model.")
        denominator = (1.0 + b_value * di * t_years) ** (1.0 / b_value)
        return [float(value) for value in qi / denominator]

    raise ValueError(f"Unsupported model name: {model_name}")


def calculate_asymptotic_exponential_curve(
    di: float,
    dates: list[datetime],
    scale_min: float = 0.0,
    scale_max: float = 1.0,
) -> list[float]:
    """Calculate asymptotic exponential curve over provided dates."""
    if not dates:
        return []

    t0 = dates[0]
    t_years = np.array([(date - t0).days / 365.25 for date in dates], dtype=float)
    q_range = scale_max - scale_min
    q_pred = scale_min + (1.0 - np.exp(-di * t_years)) * q_range
    return [float(value) for value in q_pred]


def plot_well_history(
    axis_left,
    axis_right,
    well_name: str,
    well_data: dict[str, list],
    fit_overlays: list[dict[str, Any]] | None = None,
) -> None:
    """Plot the selected well time series on a single chart with two y-axes."""
    try:
        from matplotlib.dates import DateFormatter, YearLocator
    except ImportError as exc:  # pragma: no cover - import guard for runtime use
        raise SystemExit(
            "matplotlib is required to plot the workbook data. Install dependencies "
            "with 'pip install -r requirements.txt'."
        ) from exc

    paired_series = get_sorted_well_series(well_data)

    if not paired_series:
        raise ValueError(f"No data available for well {well_name!r}.")

    sorted_dates = [item[0] for item in paired_series]
    gas_series = {
        "Qgm": [item[1] for item in paired_series],
        "Qgim": [item[6] for item in paired_series],
    }
    other_series = {
        "Qwm": [item[2] for item in paired_series],
        "Qom": [item[3] for item in paired_series],
        "Qclm": [item[4] for item in paired_series],
        "Qwim": [item[5] for item in paired_series],
    }

    axis_left.clear()
    axis_right.clear()

    gas_colors = {"Qgm": "red", "Qgim": "orange"}
    other_colors = {
        "Qwm": "blue",
        "Qom": "black",
        "Qclm": "violet",
        "Qwim": "lightblue",
    }

    for variable_name, values in gas_series.items():
        axis_left.plot(
            sorted_dates,
            values,
            marker="o",
            linewidth=1.6,
            color=gas_colors[variable_name],
            label=variable_name,
        )

    for variable_name, values in other_series.items():
        axis_right.plot(
            sorted_dates,
            values,
            marker="o",
            linewidth=1.6,
            linestyle="--",
            color=other_colors[variable_name],
            label=variable_name,
        )

    axis_left.set_ylabel("Qg / Qgi")
    axis_right.set_ylabel("Other rates")
    axis_left.set_ylim(bottom=0)
    axis_right.set_ylim(bottom=0)
    axis_left.grid(True, alpha=0.3)

    axis_left.xaxis.set_major_locator(YearLocator())
    axis_left.xaxis.set_major_formatter(DateFormatter("%Y"))
    axis_left.set_xlabel("Year")

    left_handles, left_labels = axis_left.get_legend_handles_labels()
    right_handles, right_labels = axis_right.get_legend_handles_labels()
    axis_left.legend(left_handles + right_handles, left_labels + right_labels, loc="upper left", ncol=2)

    axis_left.figure.suptitle(f"Well history: {well_name}", fontsize=14)

    if fit_overlays:
        fit_colors = ["green", "darkgreen", "teal", "brown", "magenta", "navy"]
        for fit_index, fit_overlay in enumerate(fit_overlays, start=1):
            if not fit_overlay.get("fit_dates"):
                continue

            fit_color = fit_colors[(fit_index - 1) % len(fit_colors)]
            target_axis = axis_left if fit_overlay.get("variable") == "Qgm" else axis_right
            model_name = fit_overlay["model"]
            di_value = fit_overlay["Di"]
            b_value = fit_overlay["b"]
            di_text = "n/a" if di_value is None else f"{di_value:.4f}"
            b_text = "n/a" if b_value is None else f"{b_value:.3f}"

            target_axis.plot(
                fit_overlay["fit_dates"],
                fit_overlay["fit_values"],
                color=fit_color,
                linewidth=2.4,
                label=f"Fit {fit_index} {model_name} (Di={di_text}, b={b_text})",
            )

            if fit_overlay.get("interval_start") and fit_overlay.get("interval_end"):
                axis_left.axvspan(
                    fit_overlay["interval_start"],
                    fit_overlay["interval_end"],
                    color=fit_color,
                    alpha=0.06,
                )

        left_handles, left_labels = axis_left.get_legend_handles_labels()
        right_handles, right_labels = axis_right.get_legend_handles_labels()
        axis_left.legend(left_handles + right_handles, left_labels + right_labels, loc="upper left", ncol=2)

    axis_left.figure.tight_layout()


def launch_gui(wells: dict[str, dict[str, list]], workbook_path: Path) -> None:
    """Open a GUI that allows switching the selected well at any time."""
    try:
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from matplotlib.figure import Figure
    except ImportError as exc:  # pragma: no cover - import guard for runtime use
        raise SystemExit(
            "matplotlib is required to plot the workbook data. Install dependencies "
            "with 'pip install -r requirements.txt'."
        ) from exc

    if tk is None or ttk is None:
        raise SystemExit("tkinter is required to run the interactive GUI.")

    root = tk.Tk()
    root.title("Well History Plot")
    root.geometry("1280x820")

    controls = ttk.Frame(root, padding=10)
    controls.pack(fill="x")

    ttk.Label(controls, text=f"Workbook: {workbook_path.name}").pack(side="left", padx=(0, 16))
    ttk.Label(controls, text="Well:").pack(side="left")

    well_names = sorted(wells)
    selected_well = tk.StringVar(value=well_names[0])
    well_combo = ttk.Combobox(
        controls,
        textvariable=selected_well,
        values=well_names,
        state="readonly",
        width=36,
    )
    well_combo.pack(side="left", padx=8)

    ttk.Label(controls, text="Fit variable:").pack(side="left", padx=(16, 0))
    fit_variable = tk.StringVar(value="Qg")
    variable_combo = ttk.Combobox(
        controls,
        textvariable=fit_variable,
        values=list(FIT_VARIABLE_MAP.keys()),
        state="readonly",
        width=8,
    )
    variable_combo.pack(side="left", padx=8)

    ttk.Label(controls, text="Model:").pack(side="left")
    fit_model = tk.StringVar(value="Auto")
    model_combo = ttk.Combobox(
        controls,
        textvariable=fit_model,
        values=FIT_MODEL_OPTIONS,
        state="readonly",
        width=12,
    )
    model_combo.pack(side="left", padx=8)

    ttk.Label(controls, text="Start:").pack(side="left")
    interval_start = tk.StringVar()
    start_combo = ttk.Combobox(controls, textvariable=interval_start, state="readonly", width=12)
    start_combo.pack(side="left", padx=6)

    ttk.Label(controls, text="End:").pack(side="left")
    interval_end = tk.StringVar()
    end_combo = ttk.Combobox(controls, textvariable=interval_end, state="readonly", width=12)
    end_combo.pack(side="left", padx=6)

    fit_status = tk.StringVar(value="Select one variable (Qg, Qo or Qw), interval, then click Fit Decline")
    ttk.Label(controls, textvariable=fit_status).pack(side="right")

    fig = Figure(figsize=(14, 7), dpi=100)
    axis_left = fig.add_subplot(111)
    axis_right = axis_left.twinx()

    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, root)
    toolbar.update()

    decline_results: list[dict[str, Any]] = []
    well_fit_overlays: dict[str, list[dict[str, Any]]] = {}

    def date_to_str(value: datetime) -> str:
        return value.strftime("%Y-%m-%d")

    def str_to_date(value: str) -> datetime:
        return datetime.strptime(value, "%Y-%m-%d")

    def internal_to_fit_label(variable_name: str) -> str:
        reverse_map = {internal_name: label for label, internal_name in FIT_VARIABLE_MAP.items()}
        return reverse_map.get(variable_name, variable_name)

    def fit_label_to_internal(variable_name: str) -> str | None:
        # Accept both UI labels (Qg/Qo/Qw) and legacy/internal names (Qgm/Qom/Qwm).
        cleaned = str(variable_name).strip()
        if cleaned in FIT_VARIABLE_MAP:
            return FIT_VARIABLE_MAP[cleaned]
        valid_internal_names = {name for name, _ in VARIABLES}
        if cleaned in valid_internal_names:
            return cleaned
        return None

    def export_decline_results() -> None:
        if not decline_results:
            fit_status.set("No fit results to export")
            return

        suggested_name = f"decline_fit_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if filedialog is not None:
            output_file = filedialog.asksaveasfilename(
                title="Save fit results",
                defaultextension=".xlsx",
                initialfile=suggested_name,
                filetypes=[("Excel workbook", "*.xlsx")],
                initialdir=str(workbook_path.parent),
            )
            if not output_file:
                fit_status.set("Export cancelled")
                return
            output_path = Path(output_file)
        else:
            output_path = workbook_path.parent / suggested_name

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "decline_fits"
        sheet.append([
            "well",
            "fit_variable",
            "model",
            "fit_mode",
            "qi",
            "Di",
            "b",
            "scale_min",
            "scale_max",
            "normalized",
            "rmse",
            "interval_start",
            "interval_end",
            "points_used",
        ])

        for result in decline_results:
            sheet.append([
                result["well"],
                internal_to_fit_label(result["variable"]),
                result["model"],
                result.get("fit_mode", "auto"),
                result["qi"],
                result["Di"],
                result["b"] if result["b"] is not None else "",
                result.get("scale_min", ""),
                result.get("scale_max", ""),
                result.get("normalized", ""),
                result["rmse"],
                result["interval_start"].strftime("%Y-%m-%d"),
                result["interval_end"].strftime("%Y-%m-%d"),
                len(result.get("fit_dates", [])),
            ])

        workbook.save(output_path)
        fit_status.set(f"Exported {len(decline_results)} fit(s) to {output_path.name}")

    def load_decline_results() -> None:
        if filedialog is not None:
            input_file = filedialog.askopenfilename(
                title="Open fit results",
                filetypes=[("Excel workbook", "*.xlsx")],
                initialdir=str(workbook_path.parent),
            )
            if not input_file:
                fit_status.set("Load cancelled")
                return
            input_path = Path(input_file)
        else:
            fit_status.set("File dialog unavailable to load fits")
            return

        workbook = load_workbook(input_path, data_only=True, read_only=True)
        if "decline_fits" not in workbook.sheetnames:
            fit_status.set("Sheet decline_fits not found")
            return

        sheet = workbook["decline_fits"]
        rows = sheet.iter_rows(min_row=1, values_only=True)
        try:
            header = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            fit_status.set("Fit file is empty")
            return

        index_by_name = {name: idx for idx, name in enumerate(header)}
        required_columns = ["well", "fit_variable", "model", "interval_start", "interval_end"]
        for col_name in required_columns:
            if col_name not in index_by_name:
                fit_status.set(f"Missing required column: {col_name}")
                return

        loaded_count = 0
        skipped_count = 0
        for row in rows:
            if not row:
                continue

            well_name = normalize_well_name(row[index_by_name["well"]])
            variable_label = str(row[index_by_name["fit_variable"]]).strip() if row[index_by_name["fit_variable"]] is not None else ""
            model_name = str(row[index_by_name["model"]]).strip() if row[index_by_name["model"]] is not None else ""
            di_value = row[index_by_name["Di"]] if "Di" in index_by_name else None
            interval_start_value = row[index_by_name["interval_start"]]
            interval_end_value = row[index_by_name["interval_end"]]
            b_idx = index_by_name.get("b")
            scale_min_idx = index_by_name.get("scale_min")
            scale_max_idx = index_by_name.get("scale_max")
            normalized_idx = index_by_name.get("normalized")
            qi_idx = index_by_name.get("qi")
            rmse_idx = index_by_name.get("rmse")
            fit_mode_idx = index_by_name.get("fit_mode")

            if not well_name or not variable_label or not model_name:
                skipped_count += 1
                continue
            if well_name not in wells:
                skipped_count += 1
                continue

            variable_internal = fit_label_to_internal(variable_label)
            if variable_internal is None:
                skipped_count += 1
                continue

            try:
                interval_start = parse_date(interval_start_value)
                interval_end = parse_date(interval_end_value)
            except (ValueError, TypeError):
                skipped_count += 1
                continue

            di_float = None
            if di_value not in (None, ""):
                try:
                    di_float = float(di_value)
                except (ValueError, TypeError):
                    di_float = None

            b_value = None
            if b_idx is not None and b_idx < len(row):
                raw_b = row[b_idx]
                if raw_b not in (None, ""):
                    try:
                        b_value = float(raw_b)
                    except (TypeError, ValueError):
                        skipped_count += 1
                        continue

            scale_min = 0.0
            scale_max = 1.0
            normalized = False
            if scale_min_idx is not None and scale_min_idx < len(row) and row[scale_min_idx] not in (None, ""):
                try:
                    scale_min = float(row[scale_min_idx])
                except (TypeError, ValueError):
                    scale_min = 0.0
            if scale_max_idx is not None and scale_max_idx < len(row) and row[scale_max_idx] not in (None, ""):
                try:
                    scale_max = float(row[scale_max_idx])
                except (TypeError, ValueError):
                    scale_max = 1.0
            if normalized_idx is not None and normalized_idx < len(row) and row[normalized_idx] not in (None, ""):
                normalized = str(row[normalized_idx]).strip().lower() in {"true", "1", "yes"}

            interval_dates, interval_values = get_interval_data(
                str(well_name),
                variable_internal,
                interval_start,
                interval_end,
            )
            if len(interval_dates) < 2:
                skipped_count += 1
                continue

            try:
                forced_fit = fit_decline_model(
                    interval_dates,
                    interval_values,
                    forced_model=model_name,
                )
                fit_dates = forced_fit["fit_dates"]
                fit_values = forced_fit["fit_values"]
                qi_float = forced_fit.get("qi")
                di_float = forced_fit.get("Di")
                b_value = forced_fit.get("b")
                scale_min = forced_fit.get("scale_min", scale_min)
                scale_max = forced_fit.get("scale_max", scale_max)
                normalized = forced_fit.get("normalized", normalized)
            except (ValueError, TypeError):
                skipped_count += 1
                continue

            rmse_value = 0.0
            if rmse_idx is not None and rmse_idx < len(row) and row[rmse_idx] not in (None, ""):
                try:
                    rmse_value = float(row[rmse_idx])
                except (ValueError, TypeError):
                    rmse_value = 0.0

            fit_mode_value = "auto"
            if fit_mode_idx is not None and fit_mode_idx < len(row) and row[fit_mode_idx] not in (None, ""):
                fit_mode_value = str(row[fit_mode_idx]).strip()

            result = {
                "well": well_name,
                "variable": variable_internal,
                "model": model_name,
                "fit_mode": fit_mode_value,
                "qi": None if qi_float is None else float(qi_float),
                "Di": None if di_float is None else float(di_float),
                "b": None if b_value is None else float(b_value),
                "scale_min": scale_min,
                "scale_max": scale_max,
                "normalized": normalized,
                "interval_start": interval_start,
                "interval_end": interval_end,
                "fit_dates": fit_dates,
                "fit_values": fit_values,
                "rmse": rmse_value,
            }
            decline_results.append(result)
            well_fit_overlays.setdefault(well_name, []).append(result)
            loaded_count += 1

        fit_status.set(f"Loaded {loaded_count} fit(s) from {input_path.name}; skipped {skipped_count}")
        update_plot()

    def get_interval_data(
        well_name: str,
        target_variable: str,
        start_date: datetime,
        end_date: datetime,
    ) -> tuple[list[datetime], list[float]]:
        sorted_rows = get_sorted_well_series(wells[well_name])
        variable_index = {name: index for index, (name, _) in enumerate(VARIABLES, start=1)}[target_variable]

        dates: list[datetime] = []
        values: list[float] = []
        for row in sorted_rows:
            date_value = row[0]
            rate_value = row[variable_index]
            if date_value < start_date or date_value > end_date:
                continue
            if rate_value is None:
                continue
            rate_float = float(rate_value)
            if rate_float <= 0:
                continue
            dates.append(date_value)
            values.append(rate_float)

        return dates, values

    def refresh_interval_options(*_args) -> None:
        well_name = selected_well.get()
        if well_name not in wells:
            return

        sorted_dates = [row[0] for row in get_sorted_well_series(wells[well_name])]
        unique_dates = sorted(set(sorted_dates))
        date_labels = [date_to_str(item) for item in unique_dates]
        start_combo["values"] = date_labels
        end_combo["values"] = date_labels

        if date_labels:
            interval_start.set(date_labels[0])
            interval_end.set(date_labels[-1])

    def update_plot(*_args) -> None:
        well_name = selected_well.get()
        if well_name not in wells:
            return
        overlays = well_fit_overlays.get(well_name, [])
        plot_well_history(axis_left, axis_right, well_name, wells[well_name], fit_overlays=overlays)
        canvas.draw_idle()

    def perform_fit() -> None:
        well_name = selected_well.get()
        if well_name not in wells:
            return

        try:
            start_date = str_to_date(interval_start.get())
            end_date = str_to_date(interval_end.get())
        except ValueError:
            fit_status.set("Invalid interval dates")
            return

        if end_date < start_date:
            fit_status.set("End date must be after start date")
            return

        # Fit only the user-selected variable.
        variable_label = fit_variable.get()
        target_variable = FIT_VARIABLE_MAP.get(variable_label)
        if target_variable is None:
            fit_status.set("Invalid fit variable")
            return

        interval_dates, interval_values = get_interval_data(
            well_name,
            target_variable,
            start_date,
            end_date,
        )
        if len(interval_dates) < 2:
            fit_status.set("Need at least 2 positive points in interval")
            return

        selected_model = fit_model.get().strip() or "Auto"

        if selected_model == "Asymptotic Exponential" and variable_label != "Qw":
            fit_status.set("Asymptotic Exponential is only available for Qw")
            return

        try:
            fit = fit_decline_model(interval_dates, interval_values, forced_model=selected_model)
        except ValueError as exc:
            fit_status.set(str(exc))
            return

        result = {
            "well": well_name,
            "variable": target_variable,
            "model": fit["model"],
            "fit_mode": "auto" if selected_model == "Auto" else "forced",
            "qi": None if fit["qi"] is None else float(fit["qi"]),
            "Di": None if fit["Di"] is None else float(fit["Di"]),
            "b": None if fit["b"] is None else float(fit["b"]),
            "scale_min": fit.get("scale_min"),
            "scale_max": fit.get("scale_max"),
            "normalized": fit.get("normalized", False),
            "interval_start": start_date,
            "interval_end": end_date,
            "fit_dates": fit["fit_dates"],
            "fit_values": fit["fit_values"],
            "rmse": float(fit["rmse"]),
        }
        decline_results.append(result)
        well_fit_overlays.setdefault(well_name, []).append(result)

        di_display = "n/a" if result["Di"] is None else f"{result['Di']:.4f}"
        b_display = "n/a" if result["b"] is None else f"{result['b']:.3f}"
        mode_text = "auto" if selected_model == "Auto" else f"forced:{selected_model}"
        fit_status.set(
            f"{well_name} [{variable_label}] fit #{len(well_fit_overlays[well_name])}: {result['model']} ({mode_text}) | Di={di_display} | b={b_display} | RMSE={result['rmse']:.2f}"
        )
        update_plot()

    def clear_fits_well() -> None:
        well_name = selected_well.get()
        if well_name not in wells:
            return

        removed_count = len(well_fit_overlays.get(well_name, []))
        if removed_count == 0:
            fit_status.set(f"No fits to clear for {well_name}")
            return

        well_fit_overlays.pop(well_name, None)
        decline_results[:] = [item for item in decline_results if item.get("well") != well_name]
        fit_status.set(f"Cleared {removed_count} fit(s) for {well_name}")
        update_plot()

    fit_button = ttk.Button(controls, text="Fit Decline", command=perform_fit)
    fit_button.pack(side="left", padx=(10, 0))

    clear_well_button = ttk.Button(controls, text="Clear Fits Well", command=clear_fits_well)
    clear_well_button.pack(side="left", padx=(6, 0))

    load_button = ttk.Button(controls, text="Load Fits", command=load_decline_results)
    load_button.pack(side="left", padx=(6, 0))

    export_button = ttk.Button(controls, text="Export Fits", command=export_decline_results)
    export_button.pack(side="left", padx=(6, 0))

    well_combo.bind("<<ComboboxSelected>>", refresh_interval_options)
    well_combo.bind("<<ComboboxSelected>>", update_plot, add="+")
    refresh_interval_options()
    update_plot()

    root.decline_results = decline_results
    root.mainloop()


def main() -> None:
    workbook_path = ask_workbook_path()
    print(f"Reading workbook: {workbook_path}")

    wells = read_workbook(workbook_path)
    launch_gui(wells, workbook_path)


if __name__ == "__main__":
    main()