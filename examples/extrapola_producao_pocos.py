"""
Example: Extrapolate well production decline from an Excel workbook.

The workbook is expected to contain at least the following columns:
- Poço
- Categoria
- Classe
- Fluido
- Operação
- Zona
- Início Produção
- Qg (M m3/dia)
- Qo (m3/dia)
- Qw (m3/dia)
- Di
- b

The script reads the workbook and applies decline extrapolation where
q(t) follows the hyperbolic model q = qi / (1 + b * Di * t)^(1/b), with
automatic fallback to the exponential model q(t) = qi * exp(-Di * t) when b = 0.
Time t is expressed in years. It then plots:
- flow rate versus time,
- cumulative produced volume until the production threshold is reached,
- or a combined curve for all wells of a selected fluid.

For gas and oil wells, the extrapolation stops when the flow is lower than a
closure threshold that the user can edit in the GUI (default 0 for both fluids).
"""

from __future__ import annotations

import bisect
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:  # pragma: no cover - runtime guard
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

if TYPE_CHECKING:
    import tkinter

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit(
        "matplotlib is required to plot the results. Install dependencies with "
        "'pip install -r requirements.txt'."
    ) from exc

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - import guard for runtime use
    raise SystemExit(
        "openpyxl is required to read Excel files. Install dependencies with "
        "'pip install -r requirements.txt'."
    ) from exc


COLUMN_ALIASES = {
    "Poço": ("poco", "poço", "well", "wellname", "nomepoco"),
    "Categoria": ("categoria",),
    "Classe": ("classe",),
    "Fluido": ("fluido",),
    "Operação": ("operacao", "operacão"),
    "Zona": ("zona",),
    "Início Produção": ("inicioproducao", "inicioproducao", "inicio_producao", "inicio", "startdate"),
    "Qg": ("qg", "qgmm3dia", "qgmm3d", "qgm3dia", "qgm3d"),
    "Qo": ("qo", "qom3dia", "qom3d"),
    "Qw": ("qw", "qwm3dia", "qwm3d"),
    "Constante": ("constante", "const", "fator", "ratio"),
    "Di": ("di", "di1ano", "diano", "declini", "declinio"),
    "b": ("b",),
}


EXTRAPOLATION_END_DATE: datetime = datetime(2052, 12, 31)


def normalize_header(value: object) -> str:
    """Normalize workbook headers to a comparable form."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    replacements = str.maketrans({
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
        "¹": "1",
        "²": "2",
        "³": "3",
        "⁰": "0",
    })
    text = text.translate(replacements)
    return "".join(ch for ch in text if ch.isalnum())


def ask_workbook_path() -> Path:
    """Ask the user for the workbook path at application startup."""
    if tk is not None and filedialog is not None:
        root = tk.Tk()
        root.withdraw()
        default_dir = str(Path(__file__).resolve().parent.parent)
        selected = filedialog.askopenfilename(
            title="Select Excel workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
            initialdir=default_dir,
        )
        root.destroy()

        if selected:
            return Path(selected)

    for candidate in [
        Path(__file__).resolve().parent.parent / "SMC_Reservas.xlsx",
        Path(__file__).resolve().parent.parent / "SMC_Reservas.xlsm",
    ]:
        if candidate.exists():
            return candidate

    while True:
        typed_path = input("Enter the workbook path (.xlsx/.xlsm/.xls): ").strip().strip('"')
        if not typed_path:
            print("Path cannot be empty. Try again.")
            continue

        candidate = Path(typed_path)
        if candidate.exists() and candidate.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            return candidate

        print("Invalid workbook path. Try again.")


def parse_date(value: object) -> datetime:
    """Convert cell values to a datetime object."""
    if isinstance(value, datetime):
        return value
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("Empty date")
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError as exc:
            raise ValueError(f"Unsupported date value: {value!r}") from exc
    raise ValueError(f"Unsupported date value: {value!r}")


def next_month_start(value: datetime) -> datetime:
    """Return the first day of the month following the provided date."""
    if value.month == 12:
        return datetime(value.year + 1, 1, 1)
    return datetime(value.year, value.month + 1, 1)


def month_starts_between(start_date: datetime, end_date: datetime) -> list[datetime]:
    """Build a list of month starts on/after start_date and up to end_date."""
    month_start = datetime(start_date.year, start_date.month, 1)
    if month_start < start_date:
        month_start = next_month_start(start_date)

    dates: list[datetime] = []
    current = month_start
    while current <= end_date:
        dates.append(current)
        current = next_month_start(current)
    return dates


def format_export_date(value: object) -> str:
    """Format date values for exported spreadsheets as dd/mm/yyyy."""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)


def classify_fluid(value: object) -> str:
    """Classify the fluid as gas, oil or other."""
    text = str(value or "").strip().lower()
    if "gas" in text or "gás" in text:
        return "gas"
    if "oleo" in text or "oil" in text or "óleo" in text:
        return "oil"
    return "other"


def category_class_priority(categoria: str, classe: str) -> int:
    """Return the priority rank for extrapolation selection (lower is better)."""
    categoria_norm = normalize_header(categoria)
    classe_norm = normalize_header(classe)

    if categoria_norm == "p1" and classe_norm == "pdp":
        return 0
    if categoria_norm == "p1" and classe_norm == "pdnp":
        return 1
    if categoria_norm == "p1" and classe_norm == "pud":
        return 2
    if categoria_norm == "p2":
        return 3
    if categoria_norm == "p3":
        return 4
    return 5


def infer_constant(fluid: str, qg: float, qo: float, raw_constant: float | None) -> float:
    """Infer secondary-fluid conversion factor when Constante is not explicitly provided."""
    if raw_constant is not None:
        return float(raw_constant)

    if fluid == "gas" and qg > 0.0:
        return float(qo / qg)
    if fluid == "oil" and qo > 0.0:
        return float(qg / qo)
    return 0.0


def read_workbook(
    filepath: Path,
    sheet_name: str | None = None,
    mode: str = "sequential",
) -> dict[str, dict[str, Any]]:
    """Read the well metadata workbook and group the data by well name."""
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError("The workbook is empty.")

    header_row = rows[0]
    normalized_headers = [normalize_header(cell) for cell in header_row]
    column_map: dict[str, int] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            try:
                idx = normalized_headers.index(alias)
            except ValueError:
                continue
            column_map[canonical] = idx
            break

    required_columns = {"Poço", "Início Produção", "Qg", "Qo", "Di", "b", "Fluido"}
    missing = [name for name in required_columns if name not in column_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    wells: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        if not row or not any(value is not None and str(value).strip() != "" for value in row):
            continue

        well_name = str(row[column_map["Poço"]]).strip() if column_map["Poço"] < len(row) else ""
        if not well_name:
            continue

        try:
            start_date = parse_date(row[column_map["Início Produção"]])
            qg = float(row[column_map["Qg"]])
            qo = float(row[column_map["Qo"]])
            qw = float(row[column_map["Qw"]]) if "Qw" in column_map and row[column_map["Qw"]] is not None else 0.0
            raw_constante = (
                float(row[column_map["Constante"]])
                if "Constante" in column_map and row[column_map["Constante"]] not in (None, "")
                else None
            )
            di = float(row[column_map["Di"]])
            b = float(row[column_map["b"]])
        except (TypeError, ValueError):
            continue

        fluid = classify_fluid(row[column_map["Fluido"]] if column_map["Fluido"] < len(row) else "")
        if fluid not in {"gas", "oil"}:
            continue

        constante = infer_constant(fluid, qg, qo, raw_constante)
        qi = qg if fluid == "gas" else qo

        unique_key = well_name
        duplicate_index = 2
        while unique_key in wells:
            unique_key = f"{well_name}__{duplicate_index}"
            duplicate_index += 1

        wells[unique_key] = {
            "name": well_name,
            "categoria": str(row[column_map.get("Categoria", 0)] or "") if "Categoria" in column_map else "",
            "classe": str(row[column_map.get("Classe", 0)] or "") if "Classe" in column_map else "",
            "fluido": fluid,
            "operacao": str(row[column_map.get("Operação", 0)] or "") if "Operação" in column_map else "",
            "zona": str(row[column_map.get("Zona", 0)] or "") if "Zona" in column_map else "",
            "start_date": start_date,
            "qg": qg,
            "qo": qo,
            "qw": qw,
            "constante": constante,
            "qi": qi,
            "di": di,
            "b": b,
        }

    if not wells:
        raise ValueError("No gas or oil wells were found in the workbook.")

    return resolve_duplicate_well_rows(wells, mode=mode)


def decline_rate(qi: float, di: float, b: float, t_years: np.ndarray | float) -> np.ndarray:
    """Return decline rate using hyperbolic model and exponential fallback when b is zero."""
    t_array = np.asarray(t_years, dtype=float)

    if np.isclose(b, 0.0):
        return qi * np.exp(-di * t_array)

    denominator = 1.0 + b * di * t_array
    result = np.zeros_like(t_array, dtype=float)
    valid = denominator > 0.0
    if np.any(valid):
        result[valid] = qi / np.power(denominator[valid], 1.0 / b)
    return result


def build_decline_series(well: dict[str, Any], threshold: float) -> dict[str, np.ndarray]:
    """Build the decline curve and cumulative volume for a well."""
    if all(key in well for key in ("resolved_dates", "resolved_q", "resolved_cum")):
        return {
            "dates": np.array(well["resolved_dates"], dtype=object),
            "t_years": np.array(well.get("resolved_t_years", []), dtype=float),
            "q": np.array(well["resolved_q"], dtype=float),
            "cum": np.array(well["resolved_cum"], dtype=float),
        }

    qi = float(well["qi"])
    di = float(well["di"])
    b = float(well.get("b", 0.0))
    start_date = well["start_date"]

    if qi <= 0 or qi <= threshold:
        return {
            "dates": np.array([start_date], dtype=object),
            "t_years": np.array([0.0]),
            "q": np.array([0.0]),
            "cum": np.array([0.0]),
        }

    if EXTRAPOLATION_END_DATE is not None:
        max_date = EXTRAPOLATION_END_DATE
    else:
        max_years = 50.0
        max_date = start_date + timedelta(days=max_years * 365.25)

    date_points: list[datetime] = [start_date]
    t_points: list[float] = [0.0]
    q_points: list[float] = [qi]
    cumulative_points: list[float] = [0.0]

    for current_date in month_starts_between(start_date, max_date):
        if current_date <= start_date:
            continue

        t_years = (current_date - start_date).days / 365.25
        q_value = float(decline_rate(qi, di, b, t_years))
        if q_value <= threshold:
            break

        previous_date = date_points[-1]
        previous_q = q_points[-1]
        dt_days = (current_date - previous_date).days
        cumulative_value = cumulative_points[-1] + 0.5 * (previous_q + q_value) * dt_days

        date_points.append(current_date)
        t_points.append(t_years)
        q_points.append(float(q_value))
        cumulative_points.append(float(cumulative_value))

    dates = np.array(date_points, dtype=object)
    t_years = np.array(t_points, dtype=float)
    q_values = np.array(q_points, dtype=float)
    cumulative = np.array(cumulative_points, dtype=float)

    return {
        "dates": dates,
        "t_years": t_years,
        "q": q_values,
        "cum": cumulative,
    }


def resolve_duplicate_well_rows(
    wells: dict[str, dict[str, Any]],
    mode: str = "sequential",
) -> dict[str, dict[str, Any]]:
    """Resolve duplicate well rows (same visible name and fluid) into a single effective series.

    ``mode`` controls how overlapping production intervals of the same well are combined:
    - "sequential" (default): a newer interval truncates the previous one at its start date,
      simulating field reality where only one completion is active at a time.
    - "independent": intervals are treated as independent contributions and their flow
      rates are summed, without truncating earlier intervals.
    """
    independent_mode = mode == "independent"

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    grouped_by_name: dict[str, list[dict[str, Any]]] = {}
    for well in wells.values():
        well_name = str(well["name"])
        fluid = str(well["fluido"])
        grouped.setdefault((well_name, fluid), []).append(well)
        grouped_by_name.setdefault(well_name, []).append(well)

    resolved: dict[str, dict[str, Any]] = {}
    for (well_name, fluid), rows in grouped.items():
        # A row represents a production interval that must be truncated when a
        # newer row for the same well starts, even if fluid changes. In
        # independent mode intervals are never truncated by newer rows: every
        # row keeps producing on its own until its natural decline ends, and
        # all rows for the same well/fluid are summed together.
        name_rows = grouped_by_name.get(well_name, rows)
        sorted_starts = sorted({row["start_date"] for row in name_rows})
        next_start_by_row: list[datetime | None] = []
        for row in rows:
            if independent_mode:
                next_start_by_row.append(None)
                continue
            next_start = next(
                (start for start in sorted_starts if start > row["start_date"]),
                None,
            )
            next_start_by_row.append(next_start)

        threshold = get_threshold(fluid)
        ranks = [category_class_priority(str(row.get("categoria", "")), str(row.get("classe", ""))) for row in rows]
        series_list = [build_decline_series(row, threshold) for row in rows]
        effective_end_dates: list[datetime] = []
        for series, next_start in zip(series_list, next_start_by_row):
            series_end = series["dates"][-1]
            if next_start is None:
                effective_end_dates.append(series_end)
            else:
                effective_end_dates.append(min(series_end, next_start))

        min_start = min(row["start_date"] for row in rows)
        max_end = max(effective_end_dates)
        common_dates = month_starts_between(min_start, max_end)
        if not common_dates:
            common_dates = [min_start]
        elif common_dates[0] != min_start:
            common_dates = [min_start] + common_dates
        if not common_dates:
            continue

        selected_q: list[float] = []
        selected_categoria: list[str] = []
        selected_classe: list[str] = []
        selected_zona: list[str] = []
        selected_constante: list[float] = []

        for date_value in common_dates:
            best_q = 0.0
            best_idx: int | None = None
            total_q = 0.0

            for idx, row in enumerate(rows):
                start_date = row["start_date"]
                if date_value < start_date:
                    continue

                next_start = next_start_by_row[idx]
                if next_start is not None and date_value >= next_start:
                    # This interval is closed when the next change starts.
                    continue

                t_years = (date_value - start_date).days / 365.25
                q_value = float(decline_rate(float(row["qi"]), float(row["di"]), float(row.get("b", 0.0)), t_years))
                if q_value <= threshold:
                    q_value = 0.0

                total_q += q_value

                if q_value > best_q + 1e-12:
                    best_q = q_value
                    best_idx = idx
                elif abs(q_value - best_q) <= 1e-12 and q_value > 0.0 and best_idx is not None:
                    if ranks[idx] < ranks[best_idx]:
                        best_idx = idx

            # In independent mode, overlapping intervals add up instead of the
            # highest one winning; the dominant interval still provides the
            # categoria/classe/zona/constante label for that date.
            selected_q.append(float(total_q) if independent_mode else float(best_q))
            if best_idx is None:
                selected_categoria.append("")
                selected_classe.append("")
                selected_zona.append("")
                selected_constante.append(0.0)
            else:
                selected_categoria.append(str(rows[best_idx].get("categoria", "")))
                selected_classe.append(str(rows[best_idx].get("classe", "")))
                selected_zona.append(str(rows[best_idx].get("zona", "")))
                selected_constante.append(float(rows[best_idx].get("constante", 0.0)))

        cumulative_points: list[float] = [0.0]
        for i in range(1, len(common_dates)):
            dt_days = (common_dates[i] - common_dates[i - 1]).days
            if independent_mode:
                # Independent intervals decline smoothly, so trapezoidal
                # integration is more accurate than a step approximation.
                cumulative_value = cumulative_points[-1] + 0.5 * (selected_q[i - 1] + selected_q[i]) * dt_days
            else:
                # Use left-point integration for resolved duplicate series so step
                # changes (for example, shut-in to re-open) do not create phantom
                # production before the effective change date.
                cumulative_value = cumulative_points[-1] + selected_q[i - 1] * dt_days
            cumulative_points.append(float(cumulative_value))

        min_rank_index = min(range(len(rows)), key=lambda index: ranks[index])
        representative = rows[min_rank_index]
        all_categories = " | ".join(
            sorted({str(row.get("categoria", "")).strip() for row in rows if str(row.get("categoria", "")).strip()})
        )
        all_classes = " | ".join(
            sorted({str(row.get("classe", "")).strip() for row in rows if str(row.get("classe", "")).strip()})
        )

        t_years_points = np.array(
            [((date_value - common_dates[0]).days / 365.25) for date_value in common_dates],
            dtype=float,
        )
        resolved_well = {
            **representative,
            "categoria": all_categories or str(representative.get("categoria", "")),
            "classe": all_classes or str(representative.get("classe", "")),
            "start_date": common_dates[0],
            "qi": float(selected_q[0]) if selected_q else float(representative.get("qi", 0.0)),
            "resolved_dates": common_dates,
            "resolved_t_years": t_years_points.tolist(),
            "resolved_q": selected_q,
            "resolved_cum": cumulative_points,
            "resolved_categoria": selected_categoria,
            "resolved_classe": selected_classe,
            "resolved_zona": selected_zona,
            "resolved_constante": selected_constante,
        }

        key = well_name
        if key in resolved:
            key = f"{well_name}__{fluid}"
        resolved[key] = resolved_well

    return resolved


# User-configurable closure thresholds (in the workbook units), editable via the GUI.
CLOSURE_THRESHOLDS = {"gas": 0.0, "oil": 0.0}


def set_closure_thresholds(gas: float, oil: float) -> None:
    """Update the global gas/oil closure thresholds used by get_threshold."""
    CLOSURE_THRESHOLDS["gas"] = gas
    CLOSURE_THRESHOLDS["oil"] = oil


def get_threshold(fluid: str) -> float:
    """Return the production threshold by fluid in the same units as the workbook values."""
    return CLOSURE_THRESHOLDS["gas"] if fluid == "gas" else CLOSURE_THRESHOLDS["oil"]


def qi_source_label(fluid: str) -> str:
    """Return the source column name used as qi for the given fluid."""
    return "Qg (M m3/dia)" if fluid == "gas" else "Qo (m3/dia)"


def fluid_label(fluid: str) -> str:
    """Return a user-friendly fluid label for exports."""
    return "Gás" if fluid == "gas" else "Óleo"


def combine_text_field(wells: list[dict[str, Any]], field: str) -> str:
    """Combine unique text values from a field across wells."""
    return " | ".join(
        sorted({str(well.get(field, "")).strip() for well in wells if str(well.get(field, "")).strip()})
    )


def label_active_at_date(well: dict[str, Any], date_value: datetime, field: str) -> str:
    """Return the categoria/classe label a well was actively producing at a given date."""
    resolved_dates = well.get("resolved_dates") or []
    resolved_field = well.get(f"resolved_{field}") or []
    resolved_q = well.get("resolved_q") or []
    if not resolved_dates or not resolved_field or not resolved_q:
        return ""

    idx = bisect.bisect_right(resolved_dates, date_value) - 1
    if idx < 0 or idx >= len(resolved_field):
        return ""
    if float(resolved_q[idx]) <= 0.0:
        return ""
    return str(resolved_field[idx]).strip()


def combine_active_labels_at_date(wells: list[dict[str, Any]], date_value: datetime, field: str) -> str:
    """Combine categoria/classe labels only from wells actively producing at a given date."""
    labels = {label_active_at_date(well, date_value, field) for well in wells}
    labels.discard("")
    return " | ".join(sorted(labels))


def has_secondary_contribution(well: dict[str, Any]) -> bool:
    """Return whether a well has secondary-fluid contribution via Constante."""
    if abs(float(well.get("constante", 0.0))) > 0.0:
        return True

    resolved_const = well.get("resolved_constante", [])
    if isinstance(resolved_const, list):
        return any(abs(float(value)) > 0.0 for value in resolved_const)
    return False


def category_bucket(value: object) -> str:
    """Normalize category values to P1, P2, P3 buckets."""
    text = normalize_header(value)
    if text == "p1":
        return "P1"
    if text == "p2":
        return "P2"
    if text == "p3":
        return "P3"
    return ""


def class_bucket(value: object) -> str:
    """Normalize class values to PDP, PDNP, PUD, 5PRB, 6POS buckets."""
    text = normalize_header(value)
    if text == "pdp":
        return "PDP"
    if text == "pdnp":
        return "PDNP"
    if text == "pud":
        return "PUD"
    if text == "5prb":
        return "5PRB"
    if text == "6pos":
        return "6POS"
    return ""


def _build_bucket_flow_series(
    wells: list[dict[str, Any]],
    fluid: str,
    buckets: tuple[str, ...],
    field: str,
    bucket_parser: Any,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Build concatenated flow series partitioned by an active label field."""
    contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == fluid) or has_secondary_contribution(well)]
    if not contributor_wells:
        return np.array([], dtype=object), {bucket: np.array([]) for bucket in buckets}

    series_list = [build_effective_decline_series(well, fluid) for well in contributor_wells]
    valid_pairs = [(well, series) for well, series in zip(contributor_wells, series_list) if len(series["dates"]) > 0]
    if not valid_pairs:
        return np.array([], dtype=object), {bucket: np.array([]) for bucket in buckets}

    min_start = min(well["start_date"] for well, _ in valid_pairs)
    max_end = max(series["dates"][-1] for _, series in valid_pairs)
    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]
    if common_dates[0] != min_start:
        common_dates = [min_start] + common_dates

    common_dates_array = np.array(common_dates, dtype=object)
    bucket_flow_sum: dict[str, np.ndarray] = {
        bucket: np.zeros(len(common_dates_array), dtype=float) for bucket in buckets
    }

    for well, series in valid_pairs:
        dates = np.array(series["dates"], dtype=object)
        q_values = np.array(series["q"], dtype=float)
        offsets = np.array([(date_value - well["start_date"]).days for date_value in dates], dtype=float)
        common_offsets = np.array(
            [(date_value - well["start_date"]).days for date_value in common_dates_array],
            dtype=float,
        )
        interpolated = np.interp(common_offsets, offsets, q_values, left=0.0, right=0.0)

        for idx, date_value in enumerate(common_dates_array):
            bucket = bucket_parser(label_active_at_date(well, date_value, field))
            if bucket:
                bucket_flow_sum[bucket][idx] += interpolated[idx]

    return common_dates_array, bucket_flow_sum


def build_category_flow_series(
    wells: list[dict[str, Any]],
    fluid: str,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Build concatenated flow series by category (P1/P2/P3) for a fluid."""
    buckets = ("P1", "P2", "P3")
    return _build_bucket_flow_series(wells, fluid, buckets, "categoria", category_bucket)


def build_class_flow_series(
    wells: list[dict[str, Any]],
    fluid: str,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Build concatenated flow series by class (PDP/PDNP/PUD/5PRB/6POS) for a fluid."""
    buckets = ("PDP", "PDNP", "PUD", "5PRB", "6POS")
    return _build_bucket_flow_series(wells, fluid, buckets, "classe", class_bucket)


def build_category_cumulative_series(
    wells: list[dict[str, Any]],
    fluid: str,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Build cumulative production series by category (P1/P2/P3) for a fluid."""
    contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == fluid) or has_secondary_contribution(well)]
    if not contributor_wells:
        return np.array([], dtype=object), {"P1": np.array([]), "P2": np.array([]), "P3": np.array([])}

    series_list = [build_effective_decline_series(well, fluid) for well in contributor_wells]
    valid_pairs = [(well, series) for well, series in zip(contributor_wells, series_list) if len(series["dates"]) > 0]
    if not valid_pairs:
        return np.array([], dtype=object), {"P1": np.array([]), "P2": np.array([]), "P3": np.array([])}

    min_start = min(well["start_date"] for well, _ in valid_pairs)
    max_end = max(series["dates"][-1] for _, series in valid_pairs)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]
    if common_dates[0] != min_start:
        common_dates = [min_start] + common_dates

    common_dates_array = np.array(common_dates, dtype=object)

    category_cumulative_sum: dict[str, np.ndarray] = {
        "P1": np.zeros(len(common_dates_array), dtype=float),
        "P2": np.zeros(len(common_dates_array), dtype=float),
        "P3": np.zeros(len(common_dates_array), dtype=float),
    }

    for well, series in valid_pairs:
        dates = np.array(series["dates"], dtype=object)
        cumulative = np.array(series["cum"], dtype=float)
        if len(dates) == 0:
            continue

        resolved_category = list(well.get("resolved_categoria", []))

        # Partition each well cumulative increment to the active category at the
        # beginning of the interval [i-1, i].
        per_bucket_cum: dict[str, np.ndarray] = {
            "P1": np.zeros(len(dates), dtype=float),
            "P2": np.zeros(len(dates), dtype=float),
            "P3": np.zeros(len(dates), dtype=float),
        }

        for i in range(1, len(dates)):
            for bucket in ("P1", "P2", "P3"):
                per_bucket_cum[bucket][i] = per_bucket_cum[bucket][i - 1]

            delta_cum = float(cumulative[i] - cumulative[i - 1])
            category_value = resolved_category[i - 1] if (i - 1) < len(resolved_category) else well.get("categoria", "")
            bucket = category_bucket(category_value)
            if bucket:
                per_bucket_cum[bucket][i] += delta_cum

        offsets = np.array([(date_value - well["start_date"]).days for date_value in dates], dtype=float)
        common_offsets = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)

        for bucket in ("P1", "P2", "P3"):
            bucket_cum = per_bucket_cum[bucket]
            interpolated = np.interp(
                common_offsets,
                offsets,
                bucket_cum,
                left=0.0,
                right=float(bucket_cum[-1]),
            )
            category_cumulative_sum[bucket] += interpolated

    return common_dates_array, category_cumulative_sum


def build_class_cumulative_series(
    wells: list[dict[str, Any]],
    fluid: str,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Build cumulative production series by class (PDP/PDNP/PUD/5PRB/6POS) for a fluid."""
    buckets = ("PDP", "PDNP", "PUD", "5PRB", "6POS")
    contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == fluid) or has_secondary_contribution(well)]
    if not contributor_wells:
        return np.array([], dtype=object), {bucket: np.array([]) for bucket in buckets}

    series_list = [build_effective_decline_series(well, fluid) for well in contributor_wells]
    valid_pairs = [(well, series) for well, series in zip(contributor_wells, series_list) if len(series["dates"]) > 0]
    if not valid_pairs:
        return np.array([], dtype=object), {bucket: np.array([]) for bucket in buckets}

    min_start = min(well["start_date"] for well, _ in valid_pairs)
    max_end = max(series["dates"][-1] for _, series in valid_pairs)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]
    if common_dates[0] != min_start:
        common_dates = [min_start] + common_dates

    common_dates_array = np.array(common_dates, dtype=object)
    class_cumulative_sum: dict[str, np.ndarray] = {
        bucket: np.zeros(len(common_dates_array), dtype=float) for bucket in buckets
    }

    for well, series in valid_pairs:
        dates = np.array(series["dates"], dtype=object)
        cumulative = np.array(series["cum"], dtype=float)
        if len(dates) == 0:
            continue

        resolved_class = list(well.get("resolved_classe", []))
        per_bucket_cum: dict[str, np.ndarray] = {
            bucket: np.zeros(len(dates), dtype=float) for bucket in buckets
        }

        for i in range(1, len(dates)):
            for bucket in buckets:
                per_bucket_cum[bucket][i] = per_bucket_cum[bucket][i - 1]

            delta_cum = float(cumulative[i] - cumulative[i - 1])
            class_value = resolved_class[i - 1] if (i - 1) < len(resolved_class) else well.get("classe", "")
            bucket = class_bucket(class_value)
            if bucket:
                per_bucket_cum[bucket][i] += delta_cum

        offsets = np.array([(date_value - well["start_date"]).days for date_value in dates], dtype=float)
        common_offsets = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)

        for bucket in buckets:
            bucket_cum = per_bucket_cum[bucket]
            interpolated = np.interp(
                common_offsets,
                offsets,
                bucket_cum,
                left=0.0,
                right=float(bucket_cum[-1]),
            )
            class_cumulative_sum[bucket] += interpolated

    return common_dates_array, class_cumulative_sum


def build_effective_decline_series(well: dict[str, Any], target_fluid: str) -> dict[str, np.ndarray]:
    """Return a well series converted to the requested fluid using Constante when needed."""
    source_fluid = str(well.get("fluido", ""))
    source_series = build_decline_series(well, get_threshold(source_fluid))

    dates = np.array(source_series["dates"], dtype=object)
    t_years = np.array(source_series["t_years"], dtype=float)
    q_source = np.array(source_series["q"], dtype=float)

    if len(dates) == 0:
        return {"dates": dates, "t_years": t_years, "q": np.array([], dtype=float), "cum": np.array([], dtype=float)}

    if source_fluid == target_fluid:
        return {
            "dates": dates,
            "t_years": t_years,
            "q": q_source,
            "cum": np.array(source_series["cum"], dtype=float),
        }

    resolved_const = list(well.get("resolved_constante", []))
    if resolved_const:
        constants = np.array([float(resolved_const[idx]) if idx < len(resolved_const) else float(resolved_const[-1]) for idx in range(len(dates))])
    else:
        constants = np.full(len(dates), float(well.get("constante", 0.0)), dtype=float)

    q_effective = q_source * constants
    cumulative = np.zeros(len(dates), dtype=float)
    use_left_integration = all(key in well for key in ("resolved_dates", "resolved_q", "resolved_cum"))
    for i in range(1, len(dates)):
        dt_days = (dates[i] - dates[i - 1]).days
        if use_left_integration:
            cumulative[i] = cumulative[i - 1] + q_effective[i - 1] * dt_days
        else:
            cumulative[i] = cumulative[i - 1] + 0.5 * (q_effective[i - 1] + q_effective[i]) * dt_days

    return {
        "dates": dates,
        "t_years": t_years,
        "q": q_effective,
        "cum": cumulative,
    }


def build_combined_cumulative_series_for_fluid(wells: list[dict[str, Any]], target_fluid: str) -> tuple[np.ndarray, np.ndarray]:
    """Aggregate cumulative volume for the target fluid including secondary-fluid conversion via Constante."""
    contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == target_fluid) or has_secondary_contribution(well)]
    if not contributor_wells:
        return np.array([], dtype=object), np.array([], dtype=float)

    series_list = [build_effective_decline_series(well, target_fluid) for well in contributor_wells]
    valid_pairs = [(well, series) for well, series in zip(contributor_wells, series_list) if len(series["dates"]) > 0]
    if not valid_pairs:
        return np.array([], dtype=object), np.array([], dtype=float)

    min_start = min(well["start_date"] for well, _ in valid_pairs)
    max_end = max(series["dates"][-1] for _, series in valid_pairs)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]
    if common_dates[0] != min_start:
        common_dates = [min_start] + common_dates

    common_dates_array = np.array(common_dates, dtype=object)
    summed_values = np.zeros(len(common_dates_array), dtype=float)

    for well, series in valid_pairs:
        offsets = np.array([(date_value - well["start_date"]).days for date_value in series["dates"]], dtype=float)
        cumulative = np.array(series["cum"], dtype=float)
        if len(offsets) == 0:
            continue

        common_offsets = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)
        interpolated = np.interp(
            common_offsets,
            offsets,
            cumulative,
            left=0.0,
            right=float(cumulative[-1]),
        )
        summed_values += interpolated

    return common_dates_array, summed_values


def build_combined_series(wells: list[dict[str, Any]]) -> tuple[np.ndarray, np.ndarray]:
    """Aggregate the decline rates for all wells of a given fluid on a common time grid."""
    if not wells:
        return np.array([], dtype=object), np.array([], dtype=float)

    series_list = [build_decline_series(well, get_threshold(well["fluido"])) for well in wells]
    min_start = min(well["start_date"] for well in wells)
    max_end = max(series["dates"][-1] for series in series_list)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]

    common_dates_array = np.array(common_dates, dtype=object)
    summed_values = np.zeros(len(common_dates_array), dtype=float)

    for well, series in zip(wells, series_list):
        thresholds = get_threshold(well["fluido"])
        offset_days = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)
        valid_mask = offset_days >= 0.0
        if not np.any(valid_mask):
            continue

        t_years = offset_days[valid_mask] / 365.25
        q_values = decline_rate(float(well["qi"]), float(well["di"]), float(well.get("b", 0.0)), t_years)
        q_values = np.where(q_values > thresholds, q_values, 0.0)
        summed_values[valid_mask] += q_values

    return common_dates_array, summed_values


def build_combined_series_for_fluid(wells: list[dict[str, Any]], target_fluid: str) -> tuple[np.ndarray, np.ndarray]:
    """Aggregate flow rate for the target fluid including secondary-fluid conversion via Constante."""
    contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == target_fluid) or has_secondary_contribution(well)]
    if not contributor_wells:
        return np.array([], dtype=object), np.array([], dtype=float)

    series_list = [build_effective_decline_series(well, target_fluid) for well in contributor_wells]
    valid_pairs = [(well, series) for well, series in zip(contributor_wells, series_list) if len(series["dates"]) > 0]
    if not valid_pairs:
        return np.array([], dtype=object), np.array([], dtype=float)

    min_start = min(well["start_date"] for well, _ in valid_pairs)
    max_end = max(series["dates"][-1] for _, series in valid_pairs)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]
    if common_dates[0] != min_start:
        common_dates = [min_start] + common_dates

    common_dates_array = np.array(common_dates, dtype=object)
    summed_values = np.zeros(len(common_dates_array), dtype=float)

    for well, series in valid_pairs:
        offsets = np.array([(date_value - well["start_date"]).days for date_value in series["dates"]], dtype=float)
        q_values = np.array(series["q"], dtype=float)
        if len(offsets) == 0:
            continue

        common_offsets = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)
        interpolated = np.interp(
            common_offsets,
            offsets,
            q_values,
            left=0.0,
            right=0.0,
        )
        summed_values += interpolated

    return common_dates_array, summed_values


def build_combined_cumulative_series(wells: list[dict[str, Any]]) -> tuple[np.ndarray, np.ndarray]:
    """Aggregate cumulative volume for all wells by calendar date on a common time grid."""
    if not wells:
        return np.array([], dtype=object), np.array([], dtype=float)

    series_list = [build_decline_series(well, get_threshold(well["fluido"])) for well in wells]
    min_start = min(well["start_date"] for well in wells)
    max_end = max(series["dates"][-1] for series in series_list)

    common_dates = month_starts_between(min_start, max_end)
    if not common_dates:
        common_dates = [min_start]

    common_dates_array = np.array(common_dates, dtype=object)
    summed_values = np.zeros(len(common_dates_array), dtype=float)

    for well, series in zip(wells, series_list):
        offsets = np.array([(date_value - well["start_date"]).days for date_value in series["dates"]], dtype=float)
        cumulative = np.array(series["cum"], dtype=float)
        if len(offsets) == 0:
            continue

        common_offsets = np.array([(date_value - well["start_date"]).days for date_value in common_dates_array], dtype=float)
        interpolated = np.interp(
            common_offsets,
            offsets,
            cumulative,
            left=0.0,
            right=float(cumulative[-1]),
        )
        summed_values += interpolated

    return common_dates_array, summed_values


def group_wells_by_name_and_fluid(wells: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    """Group well rows by visible well name and fluid."""
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for well in wells:
        key = (str(well["name"]), str(well["fluido"]))
        grouped.setdefault(key, []).append(well)
    return grouped


def combine_zone_labels(wells: list[dict[str, Any]]) -> str:
    """Combine zone values for grouped well rows."""
    zones = sorted({str(well.get("zona", "")).strip() for well in wells if str(well.get("zona", "")).strip()})
    return " | ".join(zones)


def detect_simultaneous_zone_production(wells: list[dict[str, Any]]) -> list[str]:
    """Detect overlapping production periods for different zones of the same well."""
    grouped_by_name: dict[str, list[dict[str, Any]]] = {}
    for well in wells:
        grouped_by_name.setdefault(str(well["name"]), []).append(well)

    warnings: list[str] = []
    for well_name, rows in grouped_by_name.items():
        if len(rows) < 2:
            continue

        intervals: list[tuple[datetime, datetime, str]] = []
        for row in rows:
            zone = str(row.get("zona", "")).strip() or "(sem zona)"
            series = build_decline_series(row, get_threshold(str(row["fluido"])))
            start = row["start_date"]
            end = series["dates"][-1]
            intervals.append((start, end, zone))

        for i in range(len(intervals)):
            start_i, end_i, zone_i = intervals[i]
            fluid_i = str(rows[i].get("fluido", ""))
            for j in range(i + 1, len(intervals)):
                start_j, end_j, zone_j = intervals[j]
                fluid_j = str(rows[j].get("fluido", ""))
                if fluid_i != fluid_j:
                    continue
                if zone_i == zone_j:
                    continue

                has_overlap = start_i <= end_j and start_j <= end_i
                if has_overlap:
                    overlap_start = max(start_i, start_j).strftime("%d/%m/%Y")
                    overlap_end = min(end_i, end_j).strftime("%d/%m/%Y")
                    warnings.append(
                        f"Poço {well_name}: sobreposição entre zonas {zone_i} e {zone_j} "
                        f"de {overlap_start} até {overlap_end}."
                    )

    return warnings


class ProductionDeclineApp:
    """Simple Tkinter-based GUI for plotting the decline extrapolation."""

    def __init__(self, master: tkinter.Tk, workbook_path: Path):
        self.master = master
        self.workbook_path = workbook_path
        self.extrapolation_modes = {
            "Sequencial (encerra produção anterior)": "sequential",
            "Independente (soma às produções anteriores)": "independent",
        }
        self.selected_mode_label = tk.StringVar(value="Sequencial (encerra produção anterior)")
        self.wells = read_workbook(workbook_path, mode=self.current_mode())
        self.zone_overlap_warnings = detect_simultaneous_zone_production(list(self.wells.values()))
        self.well_names = sorted({well["name"] for well in self.wells.values()})
        self.well_names = ["Todos os poços"] + self.well_names

        self.selected_well = tk.StringVar(value=self.well_names[0])
        self.selected_fluid = tk.StringVar(value="Gás")
        self.selected_plot = tk.StringVar(value="Vazão")
        self.gas_threshold_var = tk.StringVar(value="0")
        self.oil_threshold_var = tk.StringVar(value="0")

        self.master.title("Extrapolação de produção de poços")
        self.master.geometry("1100x700")
        self.master.protocol("WM_DELETE_WINDOW", self.on_close)
        self.create_widgets()

        if self.zone_overlap_warnings:
            warning_lines = "\n".join(self.zone_overlap_warnings[:10])
            if len(self.zone_overlap_warnings) > 10:
                warning_lines += f"\n... e mais {len(self.zone_overlap_warnings) - 10} ocorrência(s)."

            warning_message = (
                "Foram detectadas sobreposições de produção simultânea em zonas diferentes:\n\n"
                f"{warning_lines}"
            )
            if messagebox is not None:
                messagebox.showwarning("Aviso de consistência física", warning_message)
            else:
                print(warning_message)

        self.refresh_plot()

    def on_close(self) -> None:
        """Ensure Tk and Matplotlib resources are released on window close."""
        try:
            if hasattr(self, "figure"):
                plt.close(self.figure)
        finally:
            self.master.quit()
            self.master.destroy()

    def current_mode(self) -> str:
        """Return the internal extrapolation mode key for the selected combobox label."""
        return self.extrapolation_modes.get(self.selected_mode_label.get(), "sequential")

    def on_mode_change(self) -> None:
        """Reload the workbook applying the newly selected extrapolation mode."""
        self.wells = read_workbook(self.workbook_path, mode=self.current_mode())
        self.zone_overlap_warnings = detect_simultaneous_zone_production(list(self.wells.values()))
        self.refresh_plot()

    def on_threshold_change(self, *_: object) -> None:
        """Parse the closure threshold fields, reload the workbook and refresh the plot."""
        try:
            gas_threshold = float(self.gas_threshold_var.get().replace(",", "."))
            oil_threshold = float(self.oil_threshold_var.get().replace(",", "."))
        except ValueError:
            if messagebox is not None:
                messagebox.showerror("Valor inválido", "Informe números válidos para os limites de fechamento.")
            return

        set_closure_thresholds(gas_threshold, oil_threshold)
        self.wells = read_workbook(self.workbook_path, mode=self.current_mode())
        self.zone_overlap_warnings = detect_simultaneous_zone_production(list(self.wells.values()))
        self.refresh_plot()

    def create_widgets(self) -> None:
        controls = ttk.Frame(self.master, padding=10)
        controls.pack(fill="x")

        ttk.Label(controls, text="Poço").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        well_combo = ttk.Combobox(
            controls,
            textvariable=self.selected_well,
            values=self.well_names,
            state="readonly",
            width=25,
        )
        well_combo.grid(row=0, column=1, padx=5, pady=5)
        self.selected_well.trace_add("write", lambda *_: self.refresh_plot())

        ttk.Label(controls, text="Fluido").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        fluid_combo = ttk.Combobox(
            controls,
            textvariable=self.selected_fluid,
            values=["Gás", "Óleo"],
            state="readonly",
            width=15,
        )
        fluid_combo.grid(row=0, column=3, padx=5, pady=5)
        self.selected_fluid.trace_add("write", lambda *_: self.refresh_plot())

        ttk.Label(controls, text="Gráfico").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        plot_combo = ttk.Combobox(
            controls,
            textvariable=self.selected_plot,
            values=["Vazão", "Volume acumulado"],
            state="readonly",
            width=20,
        )
        plot_combo.grid(row=0, column=5, padx=5, pady=5)
        self.selected_plot.trace_add("write", lambda *_: self.refresh_plot())

        export_button = ttk.Button(controls, text="Exportar resultados", command=self.export_results)
        export_button.grid(row=0, column=6, padx=10, pady=5)

        ttk.Label(controls, text="Modo de extrapolação").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        mode_combo = ttk.Combobox(
            controls,
            textvariable=self.selected_mode_label,
            values=list(self.extrapolation_modes.keys()),
            state="readonly",
            width=40,
        )
        mode_combo.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.selected_mode_label.trace_add("write", lambda *_: self.on_mode_change())

        ttk.Label(controls, text="Limite fechamento gás").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        gas_threshold_entry = ttk.Entry(controls, textvariable=self.gas_threshold_var, width=10)
        gas_threshold_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        gas_threshold_entry.bind("<Return>", self.on_threshold_change)
        gas_threshold_entry.bind("<FocusOut>", self.on_threshold_change)

        ttk.Label(controls, text="Limite fechamento óleo").grid(row=2, column=2, padx=5, pady=5, sticky="w")
        oil_threshold_entry = ttk.Entry(controls, textvariable=self.oil_threshold_var, width=10)
        oil_threshold_entry.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        oil_threshold_entry.bind("<Return>", self.on_threshold_change)
        oil_threshold_entry.bind("<FocusOut>", self.on_threshold_change)

        update_thresholds_button = ttk.Button(
            controls, text="Atualizar extrapolação", command=self.on_threshold_change
        )
        update_thresholds_button.grid(row=2, column=4, padx=10, pady=5, sticky="w")

        self.figure, self.axes = plt.subplots(figsize=(10, 6))
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.master)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

    def get_selected_fluid(self) -> str | None:
        selected_fluid = self.selected_fluid.get().lower()
        if selected_fluid in {"gás", "gas"}:
            return "gas"
        if selected_fluid in {"óleo", "oleo", "oil"}:
            return "oil"
        return None

    def get_filtered_wells(self) -> list[dict[str, Any]]:
        fluid_filter = self.get_selected_fluid()
        if fluid_filter is None:
            return []

        selected_name = self.selected_well.get()
        if selected_name == "Todos os poços":
            return list(self.wells.values())

        matched_wells = [well for well in self.wells.values() if well["name"] == selected_name]
        if not matched_wells:
            return []

        return matched_wells

    def show_message(self, message: str) -> None:
        self.axes.clear()
        self.axes.text(0.5, 0.5, message, ha="center", va="center")
        self.axes.set_axis_off()
        self.canvas.draw()

    def refresh_plot(self) -> None:
        fluid_filter = self.get_selected_fluid()
        if fluid_filter is None:
            self.show_message("Selecione um fluido válido: gás ou óleo.")
            return

        wells = self.get_filtered_wells()
        if not wells:
            self.show_message("Nenhum poço encontrado para a seleção.")
            return

        self.axes.clear()
        plot_type = self.selected_plot.get()
        if plot_type == "Vazão":
            self.plot_flow(wells, fluid_filter)
        else:
            self.plot_cumulative(wells, fluid_filter)

        self.figure.tight_layout()
        self.canvas.draw()

    def plot_flow(self, wells: list[dict[str, Any]], fluid_filter: str) -> None:
        if self.selected_well.get() == "Todos os poços":
            dates, values = build_combined_series_for_fluid(wells, fluid_filter)
            if len(dates) == 0:
                self.show_message("Nenhum dado disponível para o fluido selecionado.")
                return
            self.axes.plot(dates, values, linewidth=2.0, color="tab:blue")
            self.axes.set_title("Vazão consolidada por fluido")
            self.axes.set_ylabel("Vazão (Mm3/dia)" if fluid_filter == "gas" else "Vazão (m3/dia)")
            self.axes.set_xlabel("Data")
            self.axes.grid(True, alpha=0.3)
            return

        selected_name = self.selected_well.get()
        dates, values = build_combined_series_for_fluid(wells, fluid_filter)
        if len(dates) == 0:
            self.show_message("Nenhum dado disponível para o fluido selecionado.")
            return

        self.axes.plot(dates, values, marker="o", linewidth=1.8, label=selected_name)

        self.axes.set_title("Vazão prevista por poço")
        self.axes.set_ylabel("Vazão (Mm3/dia)" if fluid_filter == "gas" else "Vazão (m3/dia)")
        self.axes.set_xlabel("Data")
        self.axes.grid(True, alpha=0.3)
        self.axes.legend(loc="best")

    def export_results(self) -> None:
        """Export well-level and fluid-concatenated flow/cumulative data to Excel."""
        if filedialog is None:
            self.show_message("Dialog de arquivo não disponível neste ambiente.")
            return

        default_name = "extrapolacao_producao_pocos.xlsx"
        default_dir = self.workbook_path.parent if self.workbook_path.exists() else Path.cwd()
        target = filedialog.asksaveasfilename(
            title="Salvar resultados da extrapolação",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(default_dir),
            initialfile=default_name,
        )
        if not target:
            return

        target_path = Path(target)
        try:
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            wells_sorted = sorted(self.wells.values(), key=lambda item: item["name"])
            self._add_well_flow_sheet(workbook, wells_sorted)
            self._add_well_cumulative_sheet(workbook, wells_sorted)
            self._add_fluid_flow_sheet(workbook, wells_sorted)
            self._add_fluid_cumulative_sheet(workbook, wells_sorted)
            self._add_category_flow_sheet(workbook, wells_sorted)
            self._add_class_flow_sheet(workbook, wells_sorted)
            self._add_category_cumulative_sheet(workbook, wells_sorted)
            self._add_class_cumulative_sheet(workbook, wells_sorted)

            workbook.save(target_path)
        except Exception as exc:  # pragma: no cover - runtime guard
            error_message = f"Falha ao exportar resultados: {exc}"
            if messagebox is not None:
                messagebox.showerror("Erro na exportação", error_message)
            else:
                self.show_message(error_message)
            return

        success_message = f"Arquivo exportado com sucesso:\n{target_path}"
        if messagebox is not None:
            messagebox.showinfo("Exportação concluída", success_message)
        else:
            self.show_message(success_message)

    def _add_well_flow_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("vazao_pocos")
        sheet.append(["Data", "Poço", "Categoria", "Classe", "Zona", "Qg", "Qo"])

        well_names = sorted({str(well["name"]) for well in wells})
        for well_name in well_names:
            grouped_wells = [well for well in wells if str(well["name"]) == well_name]
            zonas = combine_text_field(grouped_wells, "zona")

            dates_g, qg_values = build_combined_series_for_fluid(grouped_wells, "gas")
            dates_o, qo_values = build_combined_series_for_fluid(grouped_wells, "oil")

            all_dates = sorted(set(dates_g.tolist()) | set(dates_o.tolist()))
            if not all_dates:
                continue

            all_dates_array = np.array(all_dates, dtype=object)
            base_date = min(well["start_date"] for well in grouped_wells)
            target_offsets = np.array([(date_value - base_date).days for date_value in all_dates_array], dtype=float)

            if len(dates_g) > 0:
                gas_offsets = np.array([(date_value - base_date).days for date_value in dates_g], dtype=float)
                qg_aligned = np.interp(target_offsets, gas_offsets, qg_values, left=0.0, right=0.0)
            else:
                qg_aligned = np.zeros(len(all_dates_array), dtype=float)

            if len(dates_o) > 0:
                oil_offsets = np.array([(date_value - base_date).days for date_value in dates_o], dtype=float)
                qo_aligned = np.interp(target_offsets, oil_offsets, qo_values, left=0.0, right=0.0)
            else:
                qo_aligned = np.zeros(len(all_dates_array), dtype=float)

            for idx, date_value in enumerate(all_dates_array):
                categoria_at_date = combine_active_labels_at_date(grouped_wells, date_value, "categoria")
                classe_at_date = combine_active_labels_at_date(grouped_wells, date_value, "classe")
                sheet.append([
                    format_export_date(date_value),
                    well_name,
                    categoria_at_date,
                    classe_at_date,
                    zonas,
                    float(qg_aligned[idx]),
                    float(qo_aligned[idx]),
                ])

    def _add_well_cumulative_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("acumulada_pocos")
        sheet.append(["Data", "Poço", "Categoria", "Classe", "Gp", "Np"])

        well_names = sorted({str(well["name"]) for well in wells})
        for well_name in well_names:
            grouped_wells = [well for well in wells if str(well["name"]) == well_name]

            dates_g, gp_values = build_combined_cumulative_series_for_fluid(grouped_wells, "gas")
            dates_o, np_values = build_combined_cumulative_series_for_fluid(grouped_wells, "oil")

            all_dates = sorted(set(dates_g.tolist()) | set(dates_o.tolist()))
            if not all_dates:
                continue

            all_dates_array = np.array(all_dates, dtype=object)
            base_date = min(well["start_date"] for well in grouped_wells)
            target_offsets = np.array([(date_value - base_date).days for date_value in all_dates_array], dtype=float)

            if len(dates_g) > 0:
                gas_offsets = np.array([(date_value - base_date).days for date_value in dates_g], dtype=float)
                gp_aligned = np.interp(target_offsets, gas_offsets, gp_values, left=0.0, right=float(gp_values[-1]))
            else:
                gp_aligned = np.zeros(len(all_dates_array), dtype=float)

            if len(dates_o) > 0:
                oil_offsets = np.array([(date_value - base_date).days for date_value in dates_o], dtype=float)
                np_aligned = np.interp(target_offsets, oil_offsets, np_values, left=0.0, right=float(np_values[-1]))
            else:
                np_aligned = np.zeros(len(all_dates_array), dtype=float)

            for idx, date_value in enumerate(all_dates_array):
                categoria_at_date = combine_active_labels_at_date(grouped_wells, date_value, "categoria")
                classe_at_date = combine_active_labels_at_date(grouped_wells, date_value, "classe")
                sheet.append([
                    format_export_date(date_value),
                    well_name,
                    categoria_at_date,
                    classe_at_date,
                    float(gp_aligned[idx]),
                    float(np_aligned[idx]),
                ])

    def _add_fluid_flow_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("vazao_conc_fluido")
        sheet.append(["Data", "Fluido", "Categoria", "Classe", "Origem Qi", "Qi utilizado", "Vazão concatenada"])

        for fluid in ("gas", "oil"):
            dates, values = build_combined_series_for_fluid(wells, fluid)
            if len(dates) == 0:
                continue

            contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == fluid) or has_secondary_contribution(well)]
            primary_wells = [well for well in contributor_wells if str(well.get("fluido", "")) == fluid]
            qi_source = qi_source_label(fluid)
            qi_used = float(sum(float(well.get("qi", 0.0)) for well in primary_wells))
            for date_value, q_value in zip(dates, values):
                categoria_at_date = combine_active_labels_at_date(contributor_wells, date_value, "categoria")
                classe_at_date = combine_active_labels_at_date(contributor_wells, date_value, "classe")
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    categoria_at_date,
                    classe_at_date,
                    qi_source,
                    qi_used,
                    float(q_value),
                ])

    def _add_fluid_cumulative_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("acum_conc_fluido")
        sheet.append(["Data", "Fluido", "Categoria", "Classe", "Origem Qi", "Qi utilizado", "Acumulada concatenada"])

        for fluid in ("gas", "oil"):
            dates, values = build_combined_cumulative_series_for_fluid(wells, fluid)
            if len(dates) == 0:
                continue

            contributor_wells = [well for well in wells if (str(well.get("fluido", "")) == fluid) or has_secondary_contribution(well)]
            primary_wells = [well for well in contributor_wells if str(well.get("fluido", "")) == fluid]
            qi_source = qi_source_label(fluid)
            qi_used = float(sum(float(well["qi"]) for well in primary_wells))
            for date_value, cum_value in zip(dates, values):
                categoria_at_date = combine_active_labels_at_date(contributor_wells, date_value, "categoria")
                classe_at_date = combine_active_labels_at_date(contributor_wells, date_value, "classe")
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    categoria_at_date,
                    classe_at_date,
                    qi_source,
                    qi_used,
                    float(cum_value),
                ])

    def _add_category_cumulative_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("acum_conc_categoria")
        sheet.append(["Data", "Fluido", "Acumulada P1", "Acumulada P2", "Acumulada P3"])

        for fluid in ("gas", "oil"):
            dates, category_series = build_category_cumulative_series(wells, fluid)
            if len(dates) == 0:
                continue

            values_p1 = category_series.get("P1", np.zeros(len(dates), dtype=float))
            values_p2 = category_series.get("P2", np.zeros(len(dates), dtype=float))
            values_p3 = category_series.get("P3", np.zeros(len(dates), dtype=float))

            for idx, date_value in enumerate(dates):
                p1 = float(values_p1[idx]) if idx < len(values_p1) else 0.0
                p2 = float(values_p2[idx]) if idx < len(values_p2) else 0.0
                p3 = float(values_p3[idx]) if idx < len(values_p3) else 0.0
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    p1,
                    p2,
                    p3,
                ])

    def _add_category_flow_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("vazao_conc_categoria")
        sheet.append(["Data", "Fluido", "Vazão P1", "Vazão P2", "Vazão P3"])

        for fluid in ("gas", "oil"):
            dates, category_series = build_category_flow_series(wells, fluid)
            if len(dates) == 0:
                continue

            for idx, date_value in enumerate(dates):
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    float(category_series["P1"][idx]),
                    float(category_series["P2"][idx]),
                    float(category_series["P3"][idx]),
                ])

    def _add_class_flow_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("vazao_conc_classe")
        sheet.append(["Data", "Fluido", "Vazão PDP", "Vazão PDNP", "Vazão PUD", "Vazão 5PRB", "Vazão 6POS"])

        for fluid in ("gas", "oil"):
            dates, class_series = build_class_flow_series(wells, fluid)
            if len(dates) == 0:
                continue

            for idx, date_value in enumerate(dates):
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    float(class_series["PDP"][idx]),
                    float(class_series["PDNP"][idx]),
                    float(class_series["PUD"][idx]),
                    float(class_series["5PRB"][idx]),
                    float(class_series["6POS"][idx]),
                ])

    def _add_class_cumulative_sheet(self, workbook: Workbook, wells: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet("acum_conc_classe")
        sheet.append(["Data", "Fluido", "Acumulada PDP", "Acumulada PDNP", "Acumulada PUD", "Acumulada 5PRB", "Acumulada 6POS"])

        for fluid in ("gas", "oil"):
            dates, class_series = build_class_cumulative_series(wells, fluid)
            if len(dates) == 0:
                continue

            values_pdp = class_series.get("PDP", np.zeros(len(dates), dtype=float))
            values_pdnp = class_series.get("PDNP", np.zeros(len(dates), dtype=float))
            values_pud = class_series.get("PUD", np.zeros(len(dates), dtype=float))
            values_5prb = class_series.get("5PRB", np.zeros(len(dates), dtype=float))
            values_6pos = class_series.get("6POS", np.zeros(len(dates), dtype=float))

            for idx, date_value in enumerate(dates):
                pdp = float(values_pdp[idx]) if idx < len(values_pdp) else 0.0
                pdnp = float(values_pdnp[idx]) if idx < len(values_pdnp) else 0.0
                pud = float(values_pud[idx]) if idx < len(values_pud) else 0.0
                c5prb = float(values_5prb[idx]) if idx < len(values_5prb) else 0.0
                c6pos = float(values_6pos[idx]) if idx < len(values_6pos) else 0.0
                sheet.append([
                    format_export_date(date_value),
                    fluid_label(fluid),
                    pdp,
                    pdnp,
                    pud,
                    c5prb,
                    c6pos,
                ])

    def plot_cumulative(self, wells: list[dict[str, Any]], fluid_filter: str) -> None:
        if self.selected_well.get() == "Todos os poços":
            dates, values = build_combined_cumulative_series_for_fluid(list(self.wells.values()), fluid_filter)
            self.axes.plot(dates, values, linewidth=2.0, color="tab:green")
            self.axes.set_title("Volume acumulado consolidado")
            self.axes.set_ylabel("Volume acumulado (m3)" if fluid_filter == "oil" else "Volume acumulado (Mm3)")
            self.axes.set_xlabel("Data")
            self.axes.grid(True, alpha=0.3)
            return

        selected_name = self.selected_well.get()
        grouped_wells = [well for well in self.wells.values() if str(well["name"]) == selected_name]
        dates, values = build_combined_cumulative_series_for_fluid(grouped_wells, fluid_filter)
        self.axes.plot(dates, values, marker="o", linewidth=1.8, label=selected_name)

        self.axes.set_title("Volume produzido acumulado")
        self.axes.set_ylabel("Volume acumulado (m3)" if fluid_filter == "oil" else "Volume acumulado (Mm3)")
        self.axes.set_xlabel("Data")
        self.axes.grid(True, alpha=0.3)
        if len(wells) > 1:
            self.axes.legend(loc="best")


def run_gui(workbook_path: Path | None = None) -> None:
    """Launch the Tkinter-based GUI."""
    if tk is None or ttk is None:
        raise SystemExit("Tkinter is not available in this environment.")

    if workbook_path is None:
        workbook_path = ask_workbook_path()

    root = tk.Tk()
    app = ProductionDeclineApp(root, workbook_path)
    root.mainloop()


def main() -> None:
    path: Path | None = None
    headless = False

    args = sys.argv[1:]
    while args:
        arg = args.pop(0)
        if arg == "--headless":
            headless = True
        elif arg in {"-h", "--help"}:
            print("Usage: python examples/extrapola_producao_pocos.py [--headless] [workbook_path]")
            return
        else:
            candidate = Path(arg).expanduser()
            if candidate.exists():
                path = candidate
            else:
                print(f"Workbook not found: {candidate}")
                return

    if headless:
        if path is None:
            path = ask_workbook_path()
        wells = read_workbook(path)
        warnings = detect_simultaneous_zone_production(list(wells.values()))
        for warning in warnings:
            print(f"WARNING: {warning}")
        print(f"Loaded {len(wells)} wells from {path}")
        for name, well in sorted(wells.items()):
            series = build_decline_series(well, get_threshold(well["fluido"]))
            print(
                f"{name}: fluid={well['fluido']}, Qi={well['qi']}, Di={well['di']}, b={well.get('b', 0.0)}, "
                f"end_date={series['dates'][-1].date()}"
            )
        return

    run_gui(path)


if __name__ == "__main__":
    main()
